import io
import os
import sqlite3
from datetime import datetime, date, timedelta
from functools import wraps

import openpyxl
from flask import (
    Flask, render_template, request, redirect, url_for, session, flash, Response, g
)

from monotributo_data import (
    CATEGORIAS_MONOTRIBUTO, categoria_por_facturacion, get_categoria, VIGENCIA, periodo_legible
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "lavanderia.db")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-esta-clave-en-produccion")

APP_USER = os.environ.get("APP_USER", "admin")
APP_PASS = os.environ.get("APP_PASS", "admin123")

INGRESO_CATEGORIAS = ["Lavado", "Secado", "Planchado", "Combo (lavado+secado)", "Delivery", "Otro"]
GASTO_CATEGORIAS = ["Insumos", "Alquiler", "Luz", "Agua", "Gas", "Sueldos", "Mantenimiento", "Monotributo/Impuestos", "Otro"]
MEDIOS_PAGO = ["Efectivo", "Transferencia", "Tarjeta debito", "Tarjeta credito", "Billetera virtual"]
MESES_A_GENERAR_ADELANTE = 12

DIAS_ES = {0: "Lun", 1: "Mar", 2: "Mie", 3: "Jue", 4: "Vie", 5: "Sab", 6: "Dom"}
MESES_LARGO_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


def fecha_vencimiento_monotributo(year, month):
    """Vencimiento del monotributo: dia 20 del mes. Si cae sabado o domingo,
    pasa al primer dia habil siguiente (regla estandar de ARCA)."""
    d = date(year, month, 20)
    if d.weekday() == 5:  # sabado
        d += timedelta(days=2)
    elif d.weekday() == 6:  # domingo
        d += timedelta(days=1)
    return d


def fecha_corta_es(d):
    return f"{DIAS_ES[d.weekday()]} {d.day:02d}/{d.month:02d}"


# ---------- DB ----------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def _column_names(db, table):
    return {row["name"] for row in db.execute(f"PRAGMA table_info({table})").fetchall()}


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row

    db.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso','gasto')),
            fecha TEXT NOT NULL,
            categoria TEXT NOT NULL,
            monto REAL NOT NULL,
            medio_pago TEXT,
            contraparte TEXT,
            nota TEXT,
            creado_en TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS monotributo_cuenta (
            periodo TEXT PRIMARY KEY,
            categoria TEXT NOT NULL,
            rentas REAL NOT NULL DEFAULT 0,
            municipal REAL NOT NULL DEFAULT 0,
            pagado INTEGER NOT NULL DEFAULT 0,
            fecha_pago TEXT,
            observaciones TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS comprobantes_arca (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            tipo TEXT,
            punto_venta INTEGER,
            numero_desde INTEGER,
            numero_hasta INTEGER,
            receptor_doc TEXT,
            receptor_nombre TEXT,
            importe_total REAL,
            archivo_origen TEXT,
            importado_en TEXT,
            ingreso_id INTEGER,
            conciliado INTEGER NOT NULL DEFAULT 0,
            UNIQUE(tipo, punto_venta, numero_desde)
        )
    """)

    # --- Migracion: agregar columnas nuevas a transactions si faltan ---
    cols = _column_names(db, "transactions")
    nuevas_columnas = {
        "facturado": "INTEGER NOT NULL DEFAULT 0",
        "numero_factura": "TEXT",
        "en_cuotas": "INTEGER NOT NULL DEFAULT 0",
        "cantidad_cuotas": "INTEGER",
        "valor_cuota": "REAL",
    }
    for col, tipo in nuevas_columnas.items():
        if col not in cols:
            db.execute(f"ALTER TABLE transactions ADD COLUMN {col} {tipo}")

    db.commit()
    db.close()


def get_setting(key, default=None):
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key, value):
    db = get_db()
    db.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )
    db.commit()


# ---------- Auth ----------

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form.get("username", "")
        pw = request.form.get("password", "")
        if user == APP_USER and pw == APP_PASS:
            session["logged_in"] = True
            session["username"] = user
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)
        flash("Usuario o contraseña incorrectos", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------- Helpers ----------

def parse_month(month_str):
    """month_str formato YYYY-MM. Devuelve (year, month) o mes actual si invalido."""
    try:
        y, m = month_str.split("-")
        return int(y), int(m)
    except Exception:
        today = date.today()
        return today.year, today.month


def month_bounds(year, month):
    start = f"{year:04d}-{month:02d}-01"
    if month == 12:
        end = f"{year + 1:04d}-01-01"
    else:
        end = f"{year:04d}-{month + 1:02d}-01"
    return start, end


def add_months(year, month, n):
    total = (year * 12 + (month - 1)) + n
    return total // 12, total % 12 + 1


def periodo_str(year, month):
    return f"{year:04d}-{month:02d}"


# ---------- Rutas ----------

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    today = date.today()
    year, month = today.year, today.month
    start, end = month_bounds(year, month)

    ingresos_mes = db.execute(
        "SELECT COALESCE(SUM(monto),0) AS total FROM transactions WHERE tipo='ingreso' AND fecha >= ? AND fecha < ?",
        (start, end),
    ).fetchone()["total"]

    gastos_mes = db.execute(
        "SELECT COALESCE(SUM(monto),0) AS total FROM transactions WHERE tipo='gasto' AND fecha >= ? AND fecha < ?",
        (start, end),
    ).fetchone()["total"]

    # Facturacion acumulada ultimos 12 meses (para el semaforo de monotributo)
    hace_12 = today.replace(year=today.year - 1) if not (today.month == 2 and today.day == 29) else today.replace(year=today.year - 1, day=28)
    facturacion_12m = db.execute(
        "SELECT COALESCE(SUM(monto),0) AS total FROM transactions WHERE tipo='ingreso' AND fecha >= ?",
        (hace_12.isoformat(),),
    ).fetchone()["total"]

    categoria_actual = get_setting("categoria_monotributo", "B")
    cat_info = get_categoria(categoria_actual) or CATEGORIAS_MONOTRIBUTO[1]
    porcentaje_tope = (facturacion_12m / cat_info["tope_anual"] * 100) if cat_info["tope_anual"] else 0
    categoria_sugerida = categoria_por_facturacion(facturacion_12m)

    ultimos_mov = db.execute(
        "SELECT * FROM transactions ORDER BY fecha DESC, id DESC LIMIT 8"
    ).fetchall()

    # datos para grafico de los ultimos 6 meses
    periodos = []
    for i in range(5, -1, -1):
        yy, mm = year, month - i
        while mm <= 0:
            mm += 12
            yy -= 1
        periodos.append((yy, mm))

    meses_labels, meses_ingresos, meses_gastos = [], [], []
    for yy, mm in periodos:
        s, e = month_bounds(yy, mm)
        ing = db.execute(
            "SELECT COALESCE(SUM(monto),0) t FROM transactions WHERE tipo='ingreso' AND fecha >= ? AND fecha < ?",
            (s, e),
        ).fetchone()["t"]
        gas = db.execute(
            "SELECT COALESCE(SUM(monto),0) t FROM transactions WHERE tipo='gasto' AND fecha >= ? AND fecha < ?",
            (s, e),
        ).fetchone()["t"]
        meses_labels.append(f"{mm:02d}/{yy}")
        meses_ingresos.append(ing)
        meses_gastos.append(gas)

    return render_template(
        "dashboard.html",
        ingresos_mes=ingresos_mes,
        gastos_mes=gastos_mes,
        balance_mes=ingresos_mes - gastos_mes,
        facturacion_12m=facturacion_12m,
        cat_info=cat_info,
        categoria_actual=categoria_actual,
        porcentaje_tope=porcentaje_tope,
        categoria_sugerida=categoria_sugerida,
        ultimos_mov=ultimos_mov,
        meses_labels=meses_labels,
        meses_ingresos=meses_ingresos,
        meses_gastos=meses_gastos,
        vigencia=VIGENCIA,
        mes_nombre=f"{MESES_LARGO_ES[month]} {year}",
    )


@app.route("/ingresos", methods=["GET", "POST"])
@login_required
def ingresos():
    db = get_db()
    if request.method == "POST":
        try:
            monto = float(request.form.get("monto", "0").replace(",", "."))
        except ValueError:
            flash("Monto invalido", "error")
            return redirect(url_for("ingresos"))
        db.execute(
            "INSERT INTO transactions "
            "(tipo, fecha, categoria, monto, medio_pago, contraparte, nota, facturado, numero_factura, creado_en) "
            "VALUES ('ingreso', ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                request.form.get("fecha") or date.today().isoformat(),
                request.form.get("categoria"),
                monto,
                request.form.get("medio_pago"),
                request.form.get("contraparte", ""),
                request.form.get("nota", ""),
                1 if request.form.get("facturado") == "1" else 0,
                request.form.get("numero_factura", "").strip(),
                datetime.now().isoformat(),
            ),
        )
        db.commit()
        flash("Ingreso registrado", "success")
        return redirect(url_for("ingresos"))

    filas = db.execute(
        "SELECT * FROM transactions WHERE tipo='ingreso' ORDER BY fecha DESC, id DESC LIMIT 100"
    ).fetchall()
    return render_template(
        "movimientos.html",
        tipo="ingreso",
        titulo="Ingresos",
        categorias=INGRESO_CATEGORIAS,
        medios_pago=MEDIOS_PAGO,
        filas=filas,
        hoy=date.today().isoformat(),
    )


@app.route("/ingresos/<int:mov_id>/facturacion", methods=["POST"])
@login_required
def actualizar_facturacion(mov_id):
    db = get_db()
    db.execute(
        "UPDATE transactions SET facturado = ?, numero_factura = ? WHERE id = ? AND tipo = 'ingreso'",
        (
            1 if request.form.get("facturado") == "1" else 0,
            request.form.get("numero_factura", "").strip(),
            mov_id,
        ),
    )
    db.commit()
    flash("Facturacion actualizada", "success")
    return redirect(url_for("ingresos"))


@app.route("/gastos", methods=["GET", "POST"])
@login_required
def gastos():
    db = get_db()
    if request.method == "POST":
        try:
            monto = float(request.form.get("monto", "0").replace(",", "."))
        except ValueError:
            flash("Monto invalido", "error")
            return redirect(url_for("gastos"))

        medio_pago = request.form.get("medio_pago")
        en_cuotas = 1 if (medio_pago == "Tarjeta credito" and request.form.get("en_cuotas") == "1") else 0
        cantidad_cuotas = None
        valor_cuota = None
        if en_cuotas:
            try:
                cantidad_cuotas = int(request.form.get("cantidad_cuotas") or 0) or None
            except ValueError:
                cantidad_cuotas = None
            try:
                valor_cuota_raw = request.form.get("valor_cuota", "").replace(",", ".")
                valor_cuota = float(valor_cuota_raw) if valor_cuota_raw else None
            except ValueError:
                valor_cuota = None

        db.execute(
            "INSERT INTO transactions "
            "(tipo, fecha, categoria, monto, medio_pago, contraparte, nota, en_cuotas, cantidad_cuotas, valor_cuota, creado_en) "
            "VALUES ('gasto', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                request.form.get("fecha") or date.today().isoformat(),
                request.form.get("categoria"),
                monto,
                medio_pago,
                request.form.get("contraparte", ""),
                request.form.get("nota", ""),
                en_cuotas,
                cantidad_cuotas,
                valor_cuota,
                datetime.now().isoformat(),
            ),
        )
        db.commit()
        flash("Gasto registrado", "success")
        return redirect(url_for("gastos"))

    filas = db.execute(
        "SELECT * FROM transactions WHERE tipo='gasto' ORDER BY fecha DESC, id DESC LIMIT 100"
    ).fetchall()
    return render_template(
        "movimientos.html",
        tipo="gasto",
        titulo="Gastos",
        categorias=GASTO_CATEGORIAS,
        medios_pago=MEDIOS_PAGO,
        filas=filas,
        hoy=date.today().isoformat(),
    )


@app.route("/movimientos/<int:mov_id>/eliminar", methods=["POST"])
@login_required
def eliminar_movimiento(mov_id):
    db = get_db()
    row = db.execute("SELECT tipo FROM transactions WHERE id = ?", (mov_id,)).fetchone()
    db.execute("DELETE FROM transactions WHERE id = ?", (mov_id,))
    db.commit()
    flash("Movimiento eliminado", "success")
    if row and row["tipo"] == "gasto":
        return redirect(url_for("gastos"))
    return redirect(url_for("ingresos"))


@app.route("/reportes", methods=["GET"])
@login_required
def reportes():
    db = get_db()
    month_str = request.args.get("mes", date.today().strftime("%Y-%m"))
    year, month = parse_month(month_str)
    start, end = month_bounds(year, month)

    ingresos_por_cat = db.execute(
        "SELECT categoria, SUM(monto) total, COUNT(*) cant FROM transactions "
        "WHERE tipo='ingreso' AND fecha >= ? AND fecha < ? GROUP BY categoria ORDER BY total DESC",
        (start, end),
    ).fetchall()

    gastos_por_cat = db.execute(
        "SELECT categoria, SUM(monto) total, COUNT(*) cant FROM transactions "
        "WHERE tipo='gasto' AND fecha >= ? AND fecha < ? GROUP BY categoria ORDER BY total DESC",
        (start, end),
    ).fetchall()

    total_ingresos = sum(r["total"] for r in ingresos_por_cat)
    total_gastos = sum(r["total"] for r in gastos_por_cat)

    movimientos = db.execute(
        "SELECT * FROM transactions WHERE fecha >= ? AND fecha < ? ORDER BY fecha, id",
        (start, end),
    ).fetchall()

    return render_template(
        "reportes.html",
        mes_actual=f"{year:04d}-{month:02d}",
        ingresos_por_cat=ingresos_por_cat,
        gastos_por_cat=gastos_por_cat,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        balance=total_ingresos - total_gastos,
        movimientos=movimientos,
    )


@app.route("/reportes/exportar.csv")
@login_required
def exportar_csv():
    db = get_db()
    month_str = request.args.get("mes", date.today().strftime("%Y-%m"))
    year, month = parse_month(month_str)
    start, end = month_bounds(year, month)
    filas = db.execute(
        "SELECT fecha, tipo, categoria, monto, medio_pago, contraparte, nota, facturado, numero_factura, "
        "en_cuotas, cantidad_cuotas, valor_cuota FROM transactions "
        "WHERE fecha >= ? AND fecha < ? ORDER BY fecha, id",
        (start, end),
    ).fetchall()

    lines = ["fecha,tipo,categoria,monto,medio_pago,contraparte,nota,facturado,numero_factura,en_cuotas,cantidad_cuotas,valor_cuota"]
    for f in filas:
        nota = (f["nota"] or "").replace(",", ";").replace("\n", " ")
        contraparte = (f["contraparte"] or "").replace(",", ";")
        numero_factura = (f["numero_factura"] or "").replace(",", ";")
        lines.append(
            f'{f["fecha"]},{f["tipo"]},{f["categoria"]},{f["monto"]},{f["medio_pago"] or ""},{contraparte},{nota},'
            f'{"Si" if f["facturado"] else "No"},{numero_factura},'
            f'{"Si" if f["en_cuotas"] else "No"},{f["cantidad_cuotas"] or ""},{f["valor_cuota"] or ""}'
        )
    csv_data = "\n".join(lines)
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=reporte_{year:04d}-{month:02d}.csv"},
    )


@app.route("/config", methods=["GET", "POST"])
@login_required
def config():
    if request.method == "POST":
        set_setting("nombre_negocio", request.form.get("nombre_negocio", ""))
        set_setting("categoria_monotributo", request.form.get("categoria_monotributo", "B"))
        flash("Configuracion guardada", "success")
        return redirect(url_for("config"))

    return render_template(
        "config.html",
        nombre_negocio=get_setting("nombre_negocio", "Lavandería Tifón"),
        categoria_actual=get_setting("categoria_monotributo", "B"),
        categorias=CATEGORIAS_MONOTRIBUTO,
        vigencia=VIGENCIA,
    )


# ---------- Monotributo: cuenta corriente ----------

def _generar_periodos_faltantes(db):
    fecha_alta = get_setting("monotributo_fecha_alta", date.today().replace(day=1).isoformat())
    try:
        y0, m0, _ = [int(p) for p in fecha_alta.split("-")]
    except Exception:
        y0, m0 = date.today().year, date.today().month

    hoy = date.today()
    y1, m1 = add_months(hoy.year, hoy.month, MESES_A_GENERAR_ADELANTE)

    categoria_default = get_setting("categoria_monotributo", "B")
    municipal_default = float(get_setting("monotributo_municipal_default", "0") or 0)

    existentes = {row["periodo"] for row in db.execute("SELECT periodo FROM monotributo_cuenta").fetchall()}

    yy, mm = y0, m0
    while (yy, mm) <= (y1, m1):
        p = periodo_str(yy, mm)
        if p not in existentes:
            db.execute(
                "INSERT INTO monotributo_cuenta (periodo, categoria, rentas, municipal, pagado, fecha_pago, observaciones) "
                "VALUES (?, ?, 0, ?, 0, NULL, '')",
                (p, categoria_default, municipal_default),
            )
        yy, mm = add_months(yy, mm, 1)
    db.commit()


@app.route("/monotributo", methods=["GET"])
@login_required
def monotributo():
    db = get_db()
    _generar_periodos_faltantes(db)

    filas_db = db.execute(
        "SELECT * FROM monotributo_cuenta ORDER BY periodo"
    ).fetchall()

    hoy = date.today()

    filas = []
    total_pendiente = 0.0
    total_vencido = 0.0
    meses_pagados = 0
    proximo_vencimiento = None
    for f in filas_db:
        cat_info = get_categoria(f["categoria"]) or CATEGORIAS_MONOTRIBUTO[0]
        rentas = f["rentas"] or 0
        municipal = f["municipal"] or 0
        aportes = cat_info["aporte_sipa"] + cat_info["aporte_os"]
        total = cat_info["imp_integrado"] + aportes + rentas + municipal

        y, m = [int(p) for p in f["periodo"].split("-")]
        fecha_venc = fecha_vencimiento_monotributo(y, m)
        vencido = (not f["pagado"]) and fecha_venc < hoy

        if f["pagado"]:
            meses_pagados += 1
        else:
            total_pendiente += total
            if vencido:
                total_vencido += total
            if proximo_vencimiento is None:
                proximo_vencimiento = {
                    "fecha": fecha_corta_es(fecha_venc) + f"/{y}",
                    "periodo_legible": periodo_legible(f["periodo"]),
                    "total": total,
                }
        filas.append({
            "periodo": f["periodo"],
            "periodo_legible": periodo_legible(f["periodo"]),
            "vencimiento": fecha_corta_es(fecha_venc),
            "categoria": f["categoria"],
            "impositivo": cat_info["imp_integrado"],
            "aportes": aportes,
            "rentas": rentas,
            "municipal": municipal,
            "total": total,
            "pagado": f["pagado"],
            "vencido": vencido,
            "fecha_pago": f["fecha_pago"] or "",
            "observaciones": f["observaciones"] or "",
        })

    return render_template(
        "monotributo.html",
        filas=filas,
        total_pendiente=total_pendiente,
        total_vencido=total_vencido,
        meses_pagados=meses_pagados,
        proximo_vencimiento=proximo_vencimiento,
        categorias=[c["cat"] for c in CATEGORIAS_MONOTRIBUTO],
        contribuyente_nombre=get_setting("contribuyente_nombre", ""),
        contribuyente_cuit=get_setting("contribuyente_cuit", ""),
        contribuyente_lugar=get_setting("contribuyente_lugar", ""),
        municipal_default=get_setting("monotributo_municipal_default", "0"),
        fecha_alta=get_setting("monotributo_fecha_alta", date.today().replace(day=1).isoformat()),
        vigencia=VIGENCIA,
    )


@app.route("/monotributo/datos", methods=["POST"])
@login_required
def monotributo_datos():
    set_setting("contribuyente_nombre", request.form.get("contribuyente_nombre", ""))
    set_setting("contribuyente_cuit", request.form.get("contribuyente_cuit", ""))
    set_setting("contribuyente_lugar", request.form.get("contribuyente_lugar", ""))
    try:
        municipal = float((request.form.get("municipal_default") or "0").replace(",", "."))
    except ValueError:
        municipal = 0
    set_setting("monotributo_municipal_default", str(municipal))
    fecha_alta = request.form.get("fecha_alta") or date.today().replace(day=1).isoformat()
    set_setting("monotributo_fecha_alta", fecha_alta)
    flash("Datos del contribuyente actualizados", "success")
    return redirect(url_for("monotributo"))


@app.route("/monotributo/<periodo>", methods=["POST"])
@login_required
def monotributo_actualizar(periodo):
    db = get_db()
    categoria = request.form.get("categoria", "B")
    try:
        rentas = float((request.form.get("rentas") or "0").replace(",", "."))
    except ValueError:
        rentas = 0
    try:
        municipal = float((request.form.get("municipal") or "0").replace(",", "."))
    except ValueError:
        municipal = 0
    pagado = 1 if request.form.get("pagado") == "1" else 0
    fecha_pago = request.form.get("fecha_pago") or None
    observaciones = request.form.get("observaciones", "")

    db.execute(
        "UPDATE monotributo_cuenta SET categoria=?, rentas=?, municipal=?, pagado=?, fecha_pago=?, observaciones=? "
        "WHERE periodo=?",
        (categoria, rentas, municipal, pagado, fecha_pago, observaciones, periodo),
    )
    db.commit()
    return redirect(url_for("monotributo"))


# ---------- Facturacion: importar comprobantes de ARCA ----------

def _parse_numero_factura(texto):
    """'0001-00000123' -> (1, 123). Devuelve None si no matchea."""
    if not texto:
        return None
    partes = texto.strip().split("-")
    if len(partes) != 2:
        return None
    try:
        return int(partes[0]), int(partes[1])
    except ValueError:
        return None


def _formatear_numero(punto_venta, numero):
    try:
        return f"{int(punto_venta):05d}-{int(numero):08d}"
    except (TypeError, ValueError):
        return ""


def _celda_fecha(valor):
    if valor is None:
        return None
    if hasattr(valor, "date"):
        return valor.date().isoformat()
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt).date().isoformat()
        except ValueError:
            continue
    return texto or None


def _celda_num(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        # texto tipo "1.234,56" (formato AR)
        return float(str(valor).strip().replace(".", "").replace(",", "."))
    except ValueError:
        try:
            return float(valor)
        except (TypeError, ValueError):
            return None


def _celda_int(valor):
    n = _celda_num(valor)
    return int(n) if n is not None else None


def _conciliar_comprobantes(db):
    pendientes = db.execute("SELECT * FROM comprobantes_arca WHERE ingreso_id IS NULL").fetchall()
    if not pendientes:
        return

    ingresos_con_num = db.execute(
        "SELECT * FROM transactions WHERE tipo='ingreso' AND numero_factura IS NOT NULL AND numero_factura != ''"
    ).fetchall()

    for c in pendientes:
        match = None
        for ing in ingresos_con_num:
            if _parse_numero_factura(ing["numero_factura"]) == (c["punto_venta"], c["numero_desde"]):
                match = ing
                break

        if not match and c["fecha"] and c["importe_total"] is not None:
            candidatos = db.execute(
                "SELECT * FROM transactions WHERE tipo='ingreso' AND fecha = ? AND ABS(monto - ?) < 1 "
                "AND id NOT IN (SELECT ingreso_id FROM comprobantes_arca WHERE ingreso_id IS NOT NULL)",
                (c["fecha"], c["importe_total"]),
            ).fetchall()
            if len(candidatos) == 1:
                match = candidatos[0]

        if match:
            db.execute(
                "UPDATE comprobantes_arca SET ingreso_id=?, conciliado=1 WHERE id=?",
                (match["id"], c["id"]),
            )
            db.execute(
                "UPDATE transactions SET facturado=1, "
                "numero_factura = CASE WHEN numero_factura IS NULL OR numero_factura='' THEN ? ELSE numero_factura END "
                "WHERE id=?",
                (_formatear_numero(c["punto_venta"], c["numero_desde"]), match["id"]),
            )
    db.commit()


@app.route("/facturacion", methods=["GET"])
@login_required
def facturacion():
    db = get_db()
    comprobantes = db.execute(
        "SELECT * FROM comprobantes_arca ORDER BY fecha DESC, id DESC LIMIT 300"
    ).fetchall()

    total_importados = db.execute("SELECT COUNT(*) c FROM comprobantes_arca").fetchone()["c"]
    total_conciliados = db.execute("SELECT COUNT(*) c FROM comprobantes_arca WHERE conciliado=1").fetchone()["c"]

    comprobantes_sin_ingreso = db.execute(
        "SELECT * FROM comprobantes_arca WHERE ingreso_id IS NULL ORDER BY fecha DESC"
    ).fetchall()

    ingresos_sin_comprobante = db.execute(
        "SELECT * FROM transactions WHERE tipo='ingreso' AND facturado=1 "
        "AND id NOT IN (SELECT ingreso_id FROM comprobantes_arca WHERE ingreso_id IS NOT NULL) "
        "ORDER BY fecha DESC"
    ).fetchall()

    return render_template(
        "facturacion.html",
        comprobantes=comprobantes,
        total_importados=total_importados,
        total_conciliados=total_conciliados,
        comprobantes_sin_ingreso=comprobantes_sin_ingreso,
        ingresos_sin_comprobante=ingresos_sin_comprobante,
        formatear_numero=_formatear_numero,
    )


@app.route("/facturacion/importar", methods=["POST"])
@login_required
def facturacion_importar():
    db = get_db()
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        flash("Elegi un archivo .xlsx para importar", "error")
        return redirect(url_for("facturacion"))
    if not archivo.filename.lower().endswith(".xlsx"):
        flash("El archivo tiene que ser un .xlsx (exportado de ARCA)", "error")
        return redirect(url_for("facturacion"))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
    except Exception:
        flash("No se pudo leer el archivo. Verifica que sea el excel de 'Mis Comprobantes Emitidos' de ARCA.", "error")
        return redirect(url_for("facturacion"))

    ws = wb.active

    # Buscar la fila de encabezados (la que tiene 'Fecha' e 'Imp. Total')
    encabezados = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        valores = [str(v).strip() if v is not None else "" for v in row]
        if "Fecha" in valores and "Imp. Total" in valores:
            encabezados = valores
            header_row_idx = i
            break

    if not encabezados:
        flash("No se encontraron las columnas esperadas (Fecha, Imp. Total, etc). ¿Es el archivo correcto de ARCA?", "error")
        return redirect(url_for("facturacion"))

    idx = {nombre: pos for pos, nombre in enumerate(encabezados)}
    requeridas = ["Fecha", "Tipo", "Punto de Venta", "Número Desde", "Número Hasta",
                  "Tipo Doc. Receptor", "Nro. Doc. Receptor", "Denominación Receptor", "Imp. Total"]
    if not all(r in idx for r in requeridas):
        flash("Al archivo le faltan columnas esperadas del formato de ARCA.", "error")
        return redirect(url_for("facturacion"))

    nuevos = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        fecha = _celda_fecha(row[idx["Fecha"]])
        if not fecha:
            continue
        tipo = str(row[idx["Tipo"]] or "").strip()
        punto_venta = _celda_int(row[idx["Punto de Venta"]])
        numero_desde = _celda_int(row[idx["Número Desde"]])
        numero_hasta = _celda_int(row[idx["Número Hasta"]])
        receptor_doc = str(row[idx["Nro. Doc. Receptor"]] or "").strip()
        receptor_nombre = str(row[idx["Denominación Receptor"]] or "").strip()
        importe_total = _celda_num(row[idx["Imp. Total"]])

        if punto_venta is None or numero_desde is None:
            continue

        try:
            db.execute(
                "INSERT INTO comprobantes_arca "
                "(fecha, tipo, punto_venta, numero_desde, numero_hasta, receptor_doc, receptor_nombre, "
                "importe_total, archivo_origen, importado_en) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (fecha, tipo, punto_venta, numero_desde, numero_hasta, receptor_doc, receptor_nombre,
                 importe_total, archivo.filename, datetime.now().isoformat()),
            )
            nuevos += 1
        except sqlite3.IntegrityError:
            continue  # ya estaba importado (mismo tipo + punto de venta + numero)

    db.commit()
    _conciliar_comprobantes(db)

    if nuevos:
        flash(f"Se importaron {nuevos} comprobante(s) nuevo(s).", "success")
    else:
        flash("El archivo no tenia comprobantes nuevos para importar (o esta vacio).", "success")
    return redirect(url_for("facturacion"))


@app.context_processor
def inject_globals():
    return {"nombre_negocio": get_setting("nombre_negocio", "Lavandería Tifón") if session.get("logged_in") else ""}


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
